import openpyxl
from random import randint

from Config import config_dict


class AppData():
    def __init__(self, path):
        try:
            self.path = path
            self.file = openpyxl.load_workbook(path)
        except Exception as e:
            print("Open File Error: ", e)
            raise e

        self.sheet = self.file.active  # 当前活动的表格

    def SelectIndex(self, index=1):
        '''
        说明: 获取当前表的第index行,index从1开始
        args{
            index:要获取行的下标,默认第一行(如果index大于最大行数,则返回最后一行的数据)
        }
        return: 数据字典,key为数据在表中的下标, value为数据列表,与列名顺序对应
        '''
        index = index if self.sheet.max_row > index else self.sheet.max_row
        row = list(self.sheet.rows)[index - 1]
        result = list()
        for cell in row:
            result.append(cell.value)
        return result

    def SelectWhere(self, where={}, num=1, compare=1):
        '''
        说明: 附带条件的查询,使用要求为表的第一行为列名,返回的数据字典 key为该数据的下标,value为改行的数据值
        args{
            where:查询条件(条件字典,key为列名,value为列值)如果key不存在,则忽略该key
            num: 返回结果的条数,如果为-1 则返回全部
            compare: [1.条件等于] [2.条件小于] [3.条件大于]
        }
        return: 数据字典,key为数据在表中的下标, value为数据列表,与列名顺序对应
        '''
        result = {}
        columes_name = self.SelectIndex(index=1)  # 获取表中的列名
        conditions = {}
        for key in where:
            if key in columes_name:
                conditions[columes_name.index(key)] = where[key]
        index = 1
        for row in self.sheet.rows:
            if index == 1:
                index += 1
                continue
            flag = True

            for i in conditions:
                # print(row[i].value, conditions[i])
                if compare == 1 and str(row[i].value) != str(conditions[i]):  # 判断等于
                    flag = False
                    break
                elif compare == 2 and str(row[i].value) >= str(conditions[i]):  # 判断小于
                    flag = False
                    break
                elif compare == 3 and str(row[i].value) <= str(conditions[i]):  # 大于
                    flag = False
                    break
            if flag:
                if num == 0:
                    return result
                cells = list()
                for cell in row:
                    cells.append(cell.value)
                result[index] = cells
                num -= 1
            index += 1
        return result

    def SelectRand(self, num=1):
        '''
        说明: 随机选择num条记录返回
        args{
            num:最大条数
        }
        return: 数据字典,key为数据在表中的下标, value为数据列表,与列名顺序对应
        '''
        
        
        result = {}
        if num >= self.MaxRow():
            for i in range(3, self.sheet.max_row + 1):
                result[i] = self.SelectIndex(index=i)
            return result
        
        temp = [i for i in range(3, self.sheet.max_row + 1)]
        indexs = []
        while num > 0 and len(temp) > 0:
            indexs.append(temp.pop(randint(0, len(temp) - 1)))
            num -= 1
        for index in indexs:
            result[index] = self.SelectIndex(index=index)
        return result

    def IsInit(self, flag=False):
        '''
        说明: 检测是否初始化,是则返回True,否则初始化
        args{
            flag:是否强制初始化,当flag未真时,直接初始化
        }
        return: 
        '''

        all_table = self.__GetAllSheet()
        create_list = []
        for category in config_dict["all_category"]:
            if flag or "user_" + category not in all_table:
                create_list.append(category)
        if len(create_list):
            self.__UserInit(create_list)

    def ChangeSheet(self, sheet_name):
        '''
        说明: 切换工作表, 若存在则切换并返回True, 否则不切换并返回False
        args{
            sheet_name:要切换的工作表名
        }
        return: bool (是否切换成功)
        '''
        if sheet_name in self.__GetAllSheet():
            self.sheet = self.file[sheet_name]
            return True
        else:
            print(sheet_name, "不存在, 纯在:", self.__GetAllSheet())
            return False

    def MergeData(self, column_names, result):
        '''
        合并列名和数据字典为一个字典,返回一个数据字典的列表
        args{
            table_column_name:列名列表
            value: 值字典,key为数据的下标
        }
        '''
        datas = dict()
        for key in result:
            data = dict()
            for i in range(len(column_names)):
                data[column_names[i]] = result[key][i]
            datas[key] = data
        return datas

    def ChangeCellValue(self, row, col, value):
        self.sheet.cell(row, col).value = value
        self.__Save()
    
    def GetCellValue(self, row, col):
        return self.sheet.cell(row, col).value

    def MaxRow(self):
        return self.sheet.max_row - 2

    def __GetAllSheet(self):
        return self.file.sheetnames

    def __UserInit(self, user_study):
        """
        初始化用户的数据
        """
        for study in user_study:
            if self.__CopySheet(study, "user_" + study, True):
                print("复制{}到{}成功".format(study, ("user_" + study)))
            else:
                print("复制{}到{}失败".format(study, ("user_" + study)))

    def __DeleteSheet(self, sheet):
        """
        删除sheet表,如果存在的话
        argv{
            sheet:要删除的表名
        }
        """
        if sheet in self.__GetAllSheet():
            self.file.remove(self.file[sheet])
            self.__Save()
            return True
        return False

    def __CopySheet(self, from_sheet, to_sheet, replace=False):
        """
        复制表, replace=False时,当存在from_sheet并且不存在to_sheet时才进行复制,成功返回True 否则返回False
                replace=True时,当存在from_sheet并且存在to_sheet时会将to_sheet删除后在复制,成功返回True 否则返回False
        argv{
            from_sheet:表名, 从哪个表复制
            to_sheet:表名, 复制到哪个表
        }
        """
        sheets = self.__GetAllSheet()
        if from_sheet in sheets:
            if to_sheet in sheets and replace:
                self.__DeleteSheet(to_sheet)
            elif to_sheet in sheets and not replace:
                print("不复制{},因为{}已存在,并且不进行替换".format(from_sheet, to_sheet))
                return False
            copy = self.file.copy_worksheet(self.file[from_sheet])
            copy.title = to_sheet
            self.__Save()
            return True
        return False

    def __Save(self):
        """
        保存文件
        """
        self.file.save(self.path)

appdata = AppData("./appdata.xlsx")

if __name__ == "__main__":
    """ appdata.ChangeSheet("user_小学成语")
    print(appdata.sheet)
    print(appdata.ChangeCellValue(3, 1, 1)) """
    print(appdata.SelectRand() , appdata.SelectWhere({"记录": 0}))
    #print(appdata.SelectIndex(4))
    print(appdata.sheet.max_row)